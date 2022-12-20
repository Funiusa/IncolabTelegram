from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, PatternFill, Border, Side
from pathlib import Path
from os import path
from pandas import read_excel
from math import isnan


class FileXlsx:  # TODO use os for getting the path without path
    """ Class for creating new instance of xlsx with parameters """

    def __init__(self, path_to_file):
        self._path_to_file = path_to_file
        if not path.exists(path_to_file):
            Workbook().save(path_to_file)
        self._workbook = load_workbook(Path(path_to_file))
        self._rowCount = 0
        self._columnHeaderCount = 0
        self.current_weight_sum = 0.0
        self.received_items_list = []
        self.exists_elems = []
        self.nfound_elems = []
        self.header_for_newfile = None

    def rows_count_increment(self):
        self._rowCount += 1

    def get_current_rows_count(self):
        return self._rowCount

    def set_column_header_count(self, value):
        self._columnHeaderCount = value

    def get_column_header_count(self):
        return self._columnHeaderCount

    def save(self):
        """ Save all work """
        self._workbook.save(self._path_to_file)

    def close(self):
        """ Close file """
        self._workbook.close()

    def rows_count(self):
        """ Calculate how many rows in file """
        return self._workbook.active.max_row

    def last_column_count(self):
        """ Calculate how many column in file """
        for elem in self._workbook.active.values:
            return len(elem)

    def get_first_row(self):
        sheet = self._workbook.active
        try:
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                return tuple([elem for elem in row if elem is not None])
        except Exception as e:
            print(f"\nERROR: {e}")

    def get_values_from_column(self, header_name):
        """ Get values from column needed,
            reading file using pandas library,
            return all data without None element """
        try:
            column_data = read_excel(self._path_to_file).to_dict('list')[header_name]
            return list(filter(lambda val: not isnan(val), column_data))
        except KeyError:
            print(f'The value "{header_name}" does not exist in current file')
        except Exception as e:
            print(f"\nERROR: {e}")

    def get_row(self, value):
        """ Searching the row by value. Else return False """
        sheet = self._workbook.active
        try:
            for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
                if value in row:
                    return tuple([elem for elem in row if elem is not None])
            return False
        except Exception as e:
            print(f"\nERROR: {e}")

    def get_row_number(self, value):
        """ Searching the row by value. Else return False """
        sheet = self._workbook.active
        try:
            num = 2
            for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
                if value in row:
                    return num
                num += 1
            return 0
        except Exception as e:
            print(f"\nERROR: {e}")

    def cell_filling(self, row, column, value):
        """ Add element in one cell """
        sheet = self._workbook.active
        try:
            sheet.cell(row, column).value = value
        except ValueError:
            print(f"\nERROR: Cannot convert {value} to Excel. "
                  "Use row/column fillings methods instead.")
        except Exception as e:
            print(f"\nERROR: {e}")

    def wrap_cell(self, row, column):
        """ Wrap cell """
        sheet = self._workbook.active
        try:
            sheet.cell(row, column).alignment = Alignment(wrapText=True,
                                                          vertical='top',
                                                          horizontal='center',
                                                          shrink_to_fit=True)
        except Exception as e:
            print(f"\nERROR: {e}")

    def wrap_row(self, row, start_column, end_column):
        """ Wrap cells in rows. """
        try:
            for column in range(start_column, end_column):
                self.wrap_cell(row, column)
        except Exception as e:
            print(f"ERROR: {e}")

    def wrap_column(self, column, start_row, end_row):
        """ Wrap cells in columns """
        try:
            for row in range(start_row, end_row):
                self.wrap_cell(row, column)
        except Exception as e:
            print(f"ERROR: {e}")

    """ 'dashed', 'mediumDashed', 'thick', 'dotted', 
        'double', 'hair', 'mediumDashDot', 'mediumDashDotDot', 
        'dashDot', 'dashDotDot', 'slantDashDot', 'thin', 'medium' """

    def border_bottom_cell(self, row, column, style='medium'):  # TODO class
        """ Wrap cells in columns """
        sheet = self._workbook.active
        try:
            sheet.cell(row, column).border = Border(bottom=Side(border_style=style))
        except Exception as e:
            print(f"ERROR: {e}")

    def border_bottom_row(self, row, start_column, end_column, style='thin'):
        """ Wrap cells in rows. """
        try:
            for column in range(start_column, end_column + 1):
                self.border_bottom_cell(row, column, style)
        except Exception as e:
            print(f"ERROR: {e}")

    def row_filling(self, row, column_end, values):
        """
            Filling cells with values in row.
            All columns and rows must be start at least 1
        """
        try:
            for column in range(column_end):
                self.cell_filling(row, column + 1, values[column])
        except Exception as e:
            print(f"ERROR: {e}")

    def column_filling(self, column, start_row, values):
        """
            Filling cells with values in column
            All columns and rows must be start at least 1
        """
        try:
            for value in values:
                self.cell_filling(start_row, column, value)
                start_row += 1
        except Exception as e:
            print(f"ERROR: {e}")

    # Create subclass for commentscolor, patrn_type
    def comment_one(self, row, column, comment, authon='Noname'):  # TODO class
        """ Add comments on the cell """
        try:
            sheet = self._workbook.active
            sheet.cell(row, column).comment = Comment(comment, authon)
        except Exception as e:
            print(f"ERROR: {e}")

    def comment_row(self, row, first_column, end_column, comment, author='Noname'):
        """ Add comment for cells in row """
        try:
            for i in range(end_column):
                self.comment_one(row, first_column + i, comment, author)
        except Exception as e:
            print(f"ERROR: {e}")

    def comment_column(self, column, start_row, end_row, comment, author):
        """ Add comment for cells in column """
        try:
            for i in range(end_row - 1):
                self.comment_one(start_row + i, column, comment, author)
        except Exception as e:
            print(f"ERROR: {e}")

    """'darkHorizontal', 'lightUp', 'lightTrellis', 'lightDown', 
    'gray125', 'darkDown', 'lightHorizontal', 'lightVertical', 
    'solid', 'lightGray', 'darkTrellis', 'darkVertical', 'mediumGray',
     'darkGray', 'darkGrid', 'darkUp', 'lightGrid', 'gray0625'"""

    # Create subclass for colors
    def color_cell(self, row, column, color, patrn_type='solid'):  # TODO class
        """ Add background color for one cell """
        try:
            sheet = self._workbook.active
            sheet.cell(row, column).fill = PatternFill(patternType=patrn_type, fgColor=color)
        except Exception as e:
            print(f"ERROR: {e}")

    def color_row(self, row, first_column, end_column, color, patrn_type='solid'):
        """ Add background color for cells in row """
        try:
            for i in range(end_column):
                self.color_cell(row, first_column + i, color, patrn_type)
        except Exception as e:
            print(f"ERROR: {e}")

    def color_column(self, column, start_row, end_row, color, patrn_type='solid'):
        """ Add background color for cells in column """
        try:
            for i in range(end_row - 1):
                self.color_cell(start_row + i, column, color, patrn_type)
        except Exception as e:
            print(f"ERROR: {e}")

    def clear_color_cell(self, row, column):  # TODO class
        """ Clear one color cell """
        try:
            sheet = self._workbook.active
            sheet.cell(row, column).fill = PatternFill(fill_type=None)
        except Exception as e:
            print(f"ERROR: {e}")

    def clear_color_row(self, row, end_column):
        """ Clear cells in row """
        try:
            for column in range(end_column):
                self.clear_color_cell(row, column + 1)
        except Exception as e:
            print(f"ERROR: {e}")

    def clear_one_cell(self, row, column):
        """ Clear element in one cell """
        sheet = self._workbook.active
        try:
            sheet.cell(row, column).value = None
        except ValueError:
            print(f"\nERROR: Cannot convert to Excel. "
                  "Use row/column fillings methods instead.")
        except Exception as e:
            print(f"\nERROR: {e}")

    def clear_row(self, row, column_end):
        """
            Filling cells with values in row.
            All columns and rows must be start at least 1
        """
        try:
            for column in range(column_end):
                self.clear_one_cell(row, column + 1)
        except Exception as e:
            print(f"ERROR: {e}")

    def delete_several_rows(self, start, amount):
        sheet = self._workbook.active
        sheet.delete_rows(idx=start, amount=amount)

    def delete_row(self, index):
        sheet = self._workbook.active
        sheet.delete_rows(idx=index, amount=1)

    def freez_panes_header(self):
        """ Freeze """
        sheet = self._workbook.active
        try:
            sheet.freeze_panes = 'I2'
        except Exception as e:
            print(f"\nERROR: {e}")

    def row_stretch(self, columns):
        sheet = self._workbook.active
        try:
            for i in range(columns):
                sheet.column_dimensions[chr(ord('A') + i)].width = 10
        except Exception as e:
            print(f"\nERROR: {e}")
